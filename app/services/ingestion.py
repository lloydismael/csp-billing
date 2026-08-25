from __future__ import annotations

import csv
import datetime as dt
import ctypes
import logging
import os
from pathlib import Path

import duckdb

from app.config import settings
from app.models import Upload, UploadStatus

try:  # pragma: no cover - optional acceleration path
    import polars as pl

    _POLARS_AVAILABLE = True
except Exception:  # pragma: no cover - handled gracefully
    pl = None
    _POLARS_AVAILABLE = False

logger = logging.getLogger(__name__)


def _system_memory_total_mib() -> int | None:
    try:
        page_size = os.sysconf("SC_PAGE_SIZE")
        phys_pages = os.sysconf("SC_PHYS_PAGES")
        if page_size and phys_pages:
            return int((page_size * phys_pages) / (1024 * 1024))
    except Exception:
        pass

    if os.name == "nt":
        try:
            class MEMORYSTATUSEX(ctypes.Structure):
                _fields_ = [
                    ("dwLength", ctypes.c_ulong),
                    ("dwMemoryLoad", ctypes.c_ulong),
                    ("ullTotalPhys", ctypes.c_ulonglong),
                    ("ullAvailPhys", ctypes.c_ulonglong),
                    ("ullTotalPageFile", ctypes.c_ulonglong),
                    ("ullAvailPageFile", ctypes.c_ulonglong),
                    ("ullTotalVirtual", ctypes.c_ulonglong),
                    ("ullAvailVirtual", ctypes.c_ulonglong),
                    ("ullAvailExtendedVirtual", ctypes.c_ulonglong),
                ]

            memory_status = MEMORYSTATUSEX()
            memory_status.dwLength = ctypes.sizeof(MEMORYSTATUSEX)
            if ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(memory_status)):
                return int(memory_status.ullTotalPhys / (1024 * 1024))
        except Exception:
            pass

    return None


def _normalize_duckdb_memory_limit(value: str) -> str:
    if not value:
        return "2048MiB"

    text = str(value).strip()
    if not text.endswith("%"):
        return text

    try:
        pct = float(text[:-1])
    except ValueError:
        return "2048MiB"

    pct = max(1.0, min(95.0, pct))

    mem_total_mib = _system_memory_total_mib()
    if not mem_total_mib:
        return "2048MiB"

    target_mib = max(256, int(mem_total_mib * (pct / 100.0)))
    return f"{target_mib}MiB"

CSV_COLUMNS = [
    "PartnerId",
    "PartnerName",
    "CustomerId",
    "CustomerName",
    "CustomerDomainName",
    "CustomerCountry",
    "MpnId",
    "Tier2MpnId",
    "InvoiceNumber",
    "ProductId",
    "SkuId",
    "AvailabilityId",
    "SkuName",
    "ProductName",
    "PublisherName",
    "PublisherId",
    "SubscriptionDescription",
    "SubscriptionId",
    "ChargeStartDate",
    "ChargeEndDate",
    "UsageDate",
    "MeterType",
    "MeterCategory",
    "MeterId",
    "MeterSubCategory",
    "MeterName",
    "MeterRegion",
    "Unit",
    "ResourceLocation",
    "ConsumedService",
    "ResourceGroup",
    "ResourceURI",
    "ChargeType",
    "UnitPrice",
    "Quantity",
    "UnitType",
    "BillingPreTaxTotal",
    "BillingCurrency",
    "PricingPreTaxTotal",
    "PricingCurrency",
    "ServiceInfo1",
    "ServiceInfo2",
    "Tags",
    "AdditionalInfo",
    "EffectiveUnitPrice",
    "PCToBCExchangeRate",
    "PCToBCExchangeRateDate",
    "EntitlementId",
    "EntitlementDescription",
    "PartnerEarnedCreditPercentage",
    "CreditPercentage",
    "CreditType",
    "BenefitId",
    "BenefitOrderId",
    "BenefitType",
]


def _parquet_path(upload_id: int) -> Path:
    return settings.processed_dir / f"upload_{upload_id}.parquet"


def is_supported_upload_filename(filename: str | None) -> bool:
    if not filename:
        return False
    return Path(filename).suffix.lower() in {".csv", ".xlsx"}


def _xlsx_to_csv_path(upload: Upload, xlsx_path: Path) -> Path:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError("XLSX upload support requires openpyxl to be installed.") from exc

    csv_path = settings.processed_dir / f"upload_{upload.id}_converted.csv"
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path.unlink(missing_ok=True)

    workbook = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(["" if value is None else value for value in row])
    finally:
        workbook.close()

    return csv_path


def _source_csv_path(upload: Upload) -> tuple[Path, bool]:
    stored_path = Path(upload.stored_path)
    if stored_path.suffix.lower() == ".xlsx":
        return _xlsx_to_csv_path(upload, stored_path), True
    return stored_path, False


def _write_parquet_with_polars(csv_path: Path, parquet_path: Path) -> bool:
    if not _POLARS_AVAILABLE:
        return False

    try:
        # Stream the CSV and cast to strings to avoid expensive type inference; DuckDB will cast as needed later.
        lazy_frame = (
            pl.scan_csv(
                str(csv_path),
                has_header=True,
                infer_schema_length=settings.polars_infer_rows,
                ignore_errors=True,
                low_memory=True,
            )
            .with_columns(pl.all().cast(pl.Utf8))
        )

        lazy_frame.sink_parquet(
            str(parquet_path),
            compression="zstd",
            statistics=True,
            row_group_size=settings.polars_row_group_size,
            maintain_order=False,
            use_pyarrow=False,
        )
        return True
    except Exception as exc:  # pragma: no cover - best effort acceleration
        logger.warning("Polars ingestion fallback triggered for %s: %s", csv_path, exc)
        parquet_path.unlink(missing_ok=True)
        return False


def process_upload_csv(upload: Upload) -> dict:
    csv_path, cleanup_csv = _source_csv_path(upload)
    parquet_path = _parquet_path(upload.id)
    parquet_path.parent.mkdir(parents=True, exist_ok=True)
    parquet_path.unlink(missing_ok=True)
    settings.duckdb_temp_directory.mkdir(parents=True, exist_ok=True)

    def _literal(path: Path) -> str:
        # DuckDB COPY/read functions require string literals, so escape any single quotes.
        return path.as_posix().replace("'", "''")

    con = duckdb.connect(str(settings.duckdb_path), config={"threads": settings.duckdb_threads})
    try:
        con.execute(f"PRAGMA threads={settings.duckdb_threads}")
        normalized_memory_limit = _normalize_duckdb_memory_limit(settings.duckdb_memory_limit)
        con.execute(f"PRAGMA memory_limit='{normalized_memory_limit}'")
        con.execute(f"PRAGMA temp_directory='{_literal(settings.duckdb_temp_directory)}'")
        con.execute("CREATE SCHEMA IF NOT EXISTS uploads")

        if not _write_parquet_with_polars(csv_path, parquet_path):
            copy_sql = f"""
                COPY (
                    SELECT *
                    FROM read_csv_auto(
                        '{_literal(csv_path)}',
                        header=True,
                        union_by_name=True,
                        ignore_errors=True,
                        timestampformat='%Y-%m-%d',
                        sample_size=20000,
                        all_varchar=True
                    )
                ) TO '{_literal(parquet_path)}' (FORMAT 'parquet', COMPRESSION 'zstd');
            """
            con.execute(copy_sql)

        stats_query = f"""
            SELECT
                COUNT(*) AS total_rows,
                COALESCE(SUM(TRY_CAST(PricingPreTaxTotal AS DOUBLE)), 0) AS total_pricing,
                COALESCE(SUM(TRY_CAST(BillingPreTaxTotal AS DOUBLE)), 0) AS total_billing,
                MIN(TRY_CAST(COALESCE(UsageDate, ChargeStartDate) AS DATE)) AS min_usage,
                MAX(TRY_CAST(COALESCE(UsageDate, ChargeEndDate) AS DATE)) AS max_usage
            FROM read_parquet('{_literal(parquet_path)}')
        """
        stats = con.execute(stats_query).fetchone()

        con.execute(
            """
            CREATE OR REPLACE VIEW uploads.upload_{id} AS
            SELECT * FROM read_parquet('{path}')
            """.format(id=upload.id, path=_literal(parquet_path))
        )
    finally:
        con.close()
        if cleanup_csv:
            csv_path.unlink(missing_ok=True)

    return {
        "row_count": stats[0] or 0,
        "pricing_total": float(stats[1] or 0.0),
        "billing_total": float(stats[2] or 0.0),
        "usage_start": stats[3],
        "usage_end": stats[4],
        "parquet_path": str(parquet_path),
    }


def mark_upload_failed(upload: Upload, error: Exception, session) -> None:
    upload.status = UploadStatus.failed
    upload.error_message = str(error)
    upload.completed_at = dt.datetime.utcnow()
    session.add(upload)
    session.commit()


def mark_upload_completed(upload: Upload, stats: dict, session) -> None:
    upload.status = UploadStatus.completed
    upload.row_count = stats["row_count"]
    upload.pricing_pretax_total = stats["pricing_total"]
    upload.billing_pretax_total = stats["billing_total"]
    upload.completed_at = dt.datetime.utcnow()
    session.add(upload)
    session.commit()


def delete_upload(upload: Upload, session) -> None:
    csv_path = Path(upload.stored_path)
    parquet_path = _parquet_path(upload.id)
    csv_path.unlink(missing_ok=True)
    parquet_path.unlink(missing_ok=True)

    with duckdb.connect(str(settings.duckdb_path)) as con:
        con.execute(f"DROP VIEW IF EXISTS uploads.upload_{upload.id}")

    session.delete(upload)
    session.commit()

